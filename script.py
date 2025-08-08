#!/usr/bin/env python3

import sys
import random
import sqlite3
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from cellmap import CELL_MAP  # seu mapeamento '1X01' -> 'B4'


BASE_DIR      = Path(__file__).parent
DB_PATH       = BASE_DIR / "database.sqlite"
TEMPLATE_PATH = BASE_DIR / "EXPORT.xlsx"

OUTPUT_DIR    = BASE_DIR / "exports"          # ou Path("/var/www/html/exports")
OUTPUT_PATTERN = "Resumo_Kamishibai_{period}_{rand}.xlsx"

STATUS_QUERY = """
SELECT status
  FROM activity_records
 WHERE record_date LIKE ?
   AND item_id = ?
 LIMIT 1;
"""

def fetch_status(db_path: Path, date_filter: str, item_id: int):

    with sqlite3.connect(db_path, timeout=5) as conn:
        cur = conn.cursor()
        cur.execute(STATUS_QUERY, (date_filter, item_id))
        row = cur.fetchone()
    return row[0] if row else None

def fill_excel(template_path: Path, output_path: Path, period: str):
    """
    Para cada chave em CELL_MAP ('{item_id}X{DD}'), busca o status e escreve na célula.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    for key, cell in CELL_MAP.items():
        # '1X01' -> item_id_str='1', day_str='01'
        item_id_str, day_str = key.split('X')
        item_id = int(item_id_str)

        # Filtro 'YYYY-MM-DD%' (um dia específico dentro do mês)
        date_filter = f"{period}-{day_str}%"

        status = fetch_status(DB_PATH, date_filter, item_id) or ''
        ws[cell] = status

    wb.save(output_path)
    wb.close()

def main():
    if len(sys.argv) != 2:
        print("Uso: python script.py <YYYY-MM>")
        sys.exit(1)

    period = sys.argv[1]
    try:
        datetime.strptime(period, "%Y-%m")
    except ValueError:
        print("❌ Formato inválido. Use YYYY-MM, ex: 2025-08")
        sys.exit(1)

    if not TEMPLATE_PATH.exists():
        print(f"❌ Template não encontrado: {TEMPLATE_PATH}")
        sys.exit(1)

    # Garante a pasta de saída
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    rand = random.randint(10000, 99999)
    output_file = (OUTPUT_DIR / OUTPUT_PATTERN.format(period=period, rand=rand)).resolve()

    fill_excel(TEMPLATE_PATH, output_file, period)
    print(f"✅ Resumo gerado: {output_file}")

if __name__ == '__main__':
    main()
